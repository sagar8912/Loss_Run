from __future__ import annotations
import datetime
import json
import os
import re
import shutil
import tempfile
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple
from uuid import uuid4

import pandas as pd
import tiktoken
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

def json_dumps_datetime(obj, **kwargs):
    """
    JSON dumps with support for datetime/date objects.
    """
    def default_converter(o):
        if isinstance(o, (datetime.datetime, datetime.date)):
            return o.isoformat()
        return str(o)
    return json.dumps(obj, ensure_ascii=False, default=default_converter, **kwargs)

def _sanitize_llm_text(text: Any) -> Any:
    if not isinstance(text, str):
        return text
    patterns = [
        r"\bkill(?:ed|ing)?\b",
        r"\bmurder(?:ed|ing)?\b",
        r"\bhomicide\b",
        r"\bsuicide\b",
        r"\bassault\b",
        r"\brape\b",
        r"\bshoot(?:ing|s|er|ers)?\b",
        r"\bshot\b",
        r"\bgun(?:s)?\b",
        r"\bfirearm(?:s)?\b",
        r"\bstab(?:bed|bing)?\b",
        r"\bblood\b",
        r"\bgore\b",
        r"\bdeath(?:s)?\b",
        r"\bfatal(?:ity|ities)?\b",
        r"\binjur(?:y|ies)\b",
    ]
    sanitized = text
    for pattern in patterns:
        sanitized = re.sub(pattern, "[redacted]", sanitized, flags=re.IGNORECASE)
    return sanitized

# Advanced Excel/CSV chunking combining DataFrame cleaning with block-aware heuristics
#
# This module preserves the simple interface of 'get_excel_csv_chunks' while adding:
# - Support for legacy .xls via safe conversion to .xlsx
# - Optional block-wise chunking inspired by 'Excel_chunk.py' using openpyxl
# - Both sheet-level and block-level outputs with consistent JSON structure
#
# Usage (sheet-level only):
#    chunks = get_excel_csv_chunks_advanced(path)
#
# Usage (with block extraction):
#    chunks = get_excel_csv_chunks_advanced(path, use_blocks=True)
#
# Outputs mirror the original flow:
#    - chunk_number
#    - source_type ("csv" | "excel")
#    - sheet_name
#    - content
#
# Clean process:
#     1) Copy the DataFrame.
#     2) Rename "Unnamed: x" headers to "Column x".
#     3) Drop rows where all cells are empty/NaN.
#     4) Normalize repeated newlines in all cells (vectorized via string dtype).
#     5) Build a dict keyed by "Row N" containing only non-null cell values.

def clean_df_for_chunk(df: pd.DataFrame) -> dict:
    if df.empty:
        return {}
    df = df.copy()
    df.columns = [re.sub(r"^Unnamed: (\d+)$", r"Column \1", str(col)) for col in df.columns]
    df = df.dropna(how="all")
    if df.empty:
        return {}
    
    newline_re = re.compile(r"(\n\s*){2,}")
    df = df.astype("string")
    df = df.replace({newline_re: "\n"}, regex=True)
    
    columns = list(df.columns)
    data: Dict[str, Dict[str, Any]] = {}
    isna = pd.isna
    for idx, row in enumerate(df.itertuples(index=False, name=None), start=1):
        row_dict = {k: v for k, v in zip(columns, row) if not isna(v)}
        if row_dict:
            data[f"row {idx}"] = row_dict
    return data

def _clean_df_for_chunk_sheet_level_excel(df: pd.DataFrame) -> dict:
    """
    Cleans a DataFrame for sheet-level chunking from Excel.
    Keeps original dtypes (no string coercion) while normalizing headers and newlines.
    """
    df = df.copy()
    df.columns = [
        re.sub(r"^Unnamed: (\d+)$", r"Column \1", str(col))
        for col in df.columns
    ]
    df = df.dropna(how="all")
    
    df = df.map(
        lambda x: re.sub(r"(\n\s*){2,}", "\n", str(x)) if pd.notnull(x) else x
    )
    
    columns = list(df.columns)
    data: Dict[str, Dict[str, Any]] = {}
    for idx, row in enumerate(df.itertuples(index=False, name=None), start=1):
        row_dict: Dict[str, Any] = {}
        for key, val in zip(columns, row):
            if not pd.isna(val):
                row_dict[key] = val
        if row_dict:
            data[f"row {idx}"] = row_dict
    return data

# --- XLS Shim -----------------------------------------------------
# Converts .xls files to .xlsx for chunking compatibility.
def _copy_file_resilient(source: Path, dest: Path) -> None:
    r"""
    Copy a file to "dest" while handling Windows MAX_PATH limitations.
    On Windows this adds the extended-length path prefix (\\?\) when needed so
    long paths can still be opened even if the system-wide long path policy is
    disabled. Falls back to the original shutil.copy2 behaviour for other
    platforms.
    """
    src_for_copy = _as_extended_path(source)
    
    try:
        shutil.copy2(src_for_copy, dest)
        return
    except Exception as e:
        # Re-raise with context so the caller sees the original path for debugging.
        raise FileNotFoundError(
            f"Failed to copy file (len={len(str(source))}): {source} -> {dest}."
        ) from e

def _as_extended_path(p: Path) -> Path:
    """Return a Path with the Windows extended-length prefix when needed."""
    if os.name != "nt":
        return p
    s = str(p)
    if s.startswith("\\\\?\\") or s.startswith("//?/"):
        return Path(s)
    if s.startswith("\\\\"):
        return Path("\\\\?\\UNC\\" + s.lstrip("\\"))
    return Path("\\\\?\\" + s)

def _ensure_dir(path: Path) -> None:
    """Create a directory, using extended-length paths on Windows when necessary."""
    try:
        if os.name == "nt":
            resolved = path.resolve()
            if len(str(resolved)) >= 240:
                os.makedirs(str(_as_extended_path(resolved)), exist_ok=True)
                return
        path.mkdir(parents=True, exist_ok=True)
    except FileExistsError:
        # Race condition where the dir already exists
        pass

def _ensure_chunkable_copy(source: Path, long_path_threshold: int = 240) -> Tuple[Path, List[Path]]:
    """
    Ensure the given file path is usable by pandas/openpyxl by handling two cases:
    1) Legacy .xls files are converted to .xlsx.
    2) Windows long-path issues are avoided by copying to a short temp location.
    
    Returns
    -------
    prepared_path : Path
        Path that downstream chunkers should read.
    cleanup_paths : list[Path]
        Any temporary paths that should be deleted by the caller when finished.
    """
    cleanup_paths: List[Path] = []
    
    def _short_temp_path(suffix: str) -> Path:
        safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", source.stem)[:48] or "temp"
        return Path(tempfile.gettempdir()) / f"lossless_{safe_stem}_{uuid4().hex[:8]}{suffix}"
        
    suffix = source.suffix.lower()
    abs_source = source if source.is_absolute() else source.resolve()
    use_short_path = len(str(abs_source)) >= long_path_threshold
    
    working_source = source
    if use_short_path:
        # Copy the original file to a shorter temp path to avoid MAX_PATH issues on Windows.
        short_copy = _short_temp_path(suffix or ".tmp")
        if short_copy.exists():
            short_copy.unlink()
        try:
            _copy_file_resilient(abs_source, short_copy)
            working_source = short_copy
            cleanup_paths.append(short_copy)
        except PermissionError:
            # Fall back to using the original path if copying is blocked (e.g., OneDrive ACLs).
            working_source = abs_source
            use_short_path = False
            
    if suffix == ".xls":
        # Always place the converted file in a short, temp directory to avoid long-path failures.
        temp_path = _short_temp_path(".xlsx") if use_short_path else working_source.with_name(f"{working_source.stem}_temp.xlsx")
        if temp_path.exists():
            temp_path.unlink()
        xls = pd.ExcelFile(working_source, engine="xlrd")
        try:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                for sheet in xls.sheet_names:
                    xls.parse(sheet).to_excel(writer, sheet_name=sheet, index=False)
        finally:
            xls.close()
        cleanup_paths.append(temp_path)
        return temp_path, cleanup_paths
        
    # Non-xls files: if we copied to a short temp path, mark it for cleanup; otherwise use the original path.
    return working_source, cleanup_paths

# --- Block utilities (adapted from Excel_chunk.py, trimmed) ----------------
Coord = Tuple[int, int]
Block = Tuple[Tuple[int, int], Tuple[int, int], set]

# Utility functions for block detection and classification in Excel worksheets.
# Converts row and column indices to Excel A1 cell notation.
def a1(r: int, c: int) -> str:
    return f"{get_column_letter(c)}{r}"

# Finds all non-empty cell coordinates in a worksheet.
# Returns a set of coordinates for all non-empty cells in a worksheet.
def non_empty_coords(ws: Worksheet) -> set:
    coords = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                coords.add((r, c))
    return coords

# Groups non-empty cell coordinates into connected blocks based on adjacency.
# Returns a list of blocks, each defined by its top-left and bottom-right coordinates and member cells.
def connected_blocks(coords: set) -> List[Block]:
    from collections import deque
    
    coords = set(coords)
    seen = set()
    blocks: List[Block] = []
    for start in sorted(coords):
        if start in seen:
            continue
        q = deque([start])
        members = {start}
        seen.add(start)
        while q:
            r, c = q.popleft()
            for dr, dc in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                nb = (r + dr, c + dc)
                if nb in coords and nb not in seen:
                    seen.add(nb)
                    members.add(nb)
                    q.append(nb)
        rs = [r for r, _ in members]
        cs = [c for _, c in members]
        tl = (min(rs), min(cs))
        br = (max(rs), max(cs))
        blocks.append((tl, br, members))
    blocks.sort(key=lambda b: (b[0][0], b[0][1]))
    return blocks

# Classifies a block of cells into 'table', 'key_values', or 'text_lines' based on structure and content density.
def classify_block(ws: Worksheet, tl: Tuple[int, int], br: Tuple[int, int], members: set) -> str:
    r1, c1 = tl; r2, c2 = br
    rows = r2 - r1 + 1
    cols = c2 - c1 + 1
    total = rows * cols if rows and cols else 0
    density = (len(members) / total) if total else 0.0
    if cols == 2 and rows >= 2:
        text_like = 0
        non_empty_rows = 0
        for r in range(r1, r2 + 1):
            v1 = ws.cell(r, c1).value
            v2 = ws.cell(r, c2).value
            if v1 not in (None, "") or v2 not in (None, ""):
                non_empty_rows += 1
                if isinstance(v1, str) and v1.strip():
                    text_like += 1
        if non_empty_rows and (text_like / non_empty_rows) > 0.6:
            return "key_values"
    if rows >= 3 and cols >= 2 and density >= 0.5:
        return "table"
    if rows <= 2 and len(members) <= 8:
        return "text_lines"
    return "key_values"

# Calculates the number of overlapping columns between two rectangular blocks.
def _rect_overlap_cols(a_tl, a_br, b_tl, b_br) -> int:
    ac1, ac2 = a_tl[1], a_br[1]; bc1, bc2 = b_tl[1], b_br[1]
    left = max(ac1, bc1); right = min(ac2, bc2)
    return max(0, right - left + 1)

# Calculates the number of overlapping rows between two rectangular blocks.
def _rect_overlap_rows(a_tl, a_br, b_tl, b_br) -> int:
    ar1, ar2 = a_tl[0], a_br[0]; br1, br2 = b_tl[0], b_br[0]
    top = max(ar1, br1); bot = min(ar2, br2)
    return max(0, bot - top + 1)

# Calculates the vertical gap (in rows) between two rectangles. Returns 0 if they overlap.
def _rect_gap_rows(a_tl, a_br, b_tl, b_br) -> int:
    if _rect_overlap_rows(a_tl, a_br, b_tl, b_br) > 0:
        return 0
    ar2 = a_br[0]; br1 = b_tl[0]
    if ar2 < br1:
        return max(0, br1 - ar2 - 1)
    br2 = b_br[0]; ar1 = a_tl[0]
    return max(0, ar1 - br2 - 1)

# Calculates the horizontal gap (in columns) between two rectangles. Returns 0 if they overlap.
def _rect_gap_cols(a_tl, a_br, b_tl, b_br) -> int:
    if _rect_overlap_cols(a_tl, a_br, b_tl, b_br) > 0:
        return 0
    ac2 = a_br[1]; bc1 = b_tl[1]
    if ac2 < bc1:
        return max(0, bc1 - ac2 - 1)
    bc2 = b_br[1]; ac1 = a_tl[1]
    return max(0, ac1 - bc2 - 1)

# Returns the number of columns in a rectangle defined by top-left and bottom-right coordinates.
def _rect_col_count(tl, br) -> int:
    return br[1] - tl[1] + 1

# Returns the number of rows in a rectangle defined by top-left and bottom-right coordinates.
def _rect_row_count(tl, br) -> int:
    return br[0] - tl[0] + 1

# Returns the union of two rectangles defined by their top-left and bottom-right coordinates.
def _rect_union(a_tl, a_br, b_tl, b_br):
    (ar1, ac1), (ar2, ac2) = a_tl, a_br
    (br1, bc1), (br2, bc2) = b_tl, b_br
    return (min(ar1, br1), min(ac1, bc1)), (max(ar2, br2), max(ac2, bc2))

def _should_merge(
    label1: str,
    tl1: Tuple[int, int],
    br1: Tuple[int, int],
    label2: str,
    tl2: Tuple[int, int],
    br2: Tuple[int, int],
) -> bool:
    if label1 != label2:
        return False
    # Merge if they overlap in one dimension and are close in the other dimension
    if _rect_overlap_cols(tl1, br1, tl2, br2) > 0 and _rect_gap_rows(tl1, br1, tl2, br2) <= 1:
        return True
    if _rect_overlap_rows(tl1, br1, tl2, br2) > 0 and _rect_gap_cols(tl1, br1, tl2, br2) <= 1:
        return True
    return False

# Merges adjacent blocks based on their labels and spatial relationship.
def merge_adjacent_blocks(blocks, labels):
    merged = []
    merged_labels = []
    used = [False] * len(blocks)
    for i, blk in enumerate(blocks):
        if used[i]:
            continue
        acc_tl, acc_br = blk[0], blk[1]
        acc_label = labels[i]
        acc_members = set(blk[2])
        for j in range(i + 1, len(blocks)):
            if used[j]:
                continue
            if _should_merge(acc_label, acc_tl, acc_br, labels[j], blocks[j][0], blocks[j][1]):
                acc_tl, acc_br = _rect_union(acc_tl, acc_br, blocks[j][0], blocks[j][1])
                acc_members.update(blocks[j][2])
                used[j] = True
        merged.append((acc_tl, acc_br, acc_members))
        merged_labels.append(acc_label)
        used[i] = True
    return merged, merged_labels

# Expands table blocks by merging nearby sparse blocks (key_values/text_lines) that share layout.
# The vertical gap is intentionally lenient to allow tables split by subtotal rows or repeated headers.
def _merge_tables_with_neighbors(
    blocks,
    labels,
    max_vertical_gap: int = 6,
    max_horizontal_gap: int = 2,
):
    blocks = list(blocks)
    labels = list(labels)
    
    def _row_overlap_ratio(a_tl, a_br, b_tl, b_br) -> float:
        overlap = _rect_overlap_rows(a_tl, a_br, b_tl, b_br)
        denom = min(_rect_row_count(a_tl, a_br), _rect_row_count(b_tl, b_br))
        return (overlap / denom) if denom else 0.0
        
    def _col_overlap_ratio(a_tl, a_br, b_tl, b_br) -> float:
        overlap = _rect_overlap_cols(a_tl, a_br, b_tl, b_br)
        denom = min(_rect_col_count(a_tl, a_br), _rect_col_count(b_tl, b_br))
        return (overlap / denom) if denom else 0.0
        
    changed = True
    while changed:
        changed = False
        to_drop = set()
        for i, blk in enumerate(blocks):
            if blk is None or labels[i] != "table":
                continue
            tl_i, br_i, members_i = blk
            for j, other in enumerate(blocks):
                if i == j or other is None or j in to_drop:
                    continue
                lbl_j = labels[j]
                if lbl_j not in ("table", "key_values", "text_lines"):
                    continue
                    
                tl_j, br_j, members_j = other
                col_ratio = _col_overlap_ratio(tl_i, br_i, tl_j, br_j)
                row_ratio = _row_overlap_ratio(tl_i, br_i, tl_j, br_j)
                vgap = _rect_gap_rows(tl_i, br_i, tl_j, br_j)
                hgap = _rect_gap_cols(tl_i, br_i, tl_j, br_j)
                
                is_above = tl_j[0] < tl_i[0]
                if is_above and tl_j[0] < (tl_i[0] - max_vertical_gap):
                    continue
                near_vertical = col_ratio >= 0.5 and vgap <= max_vertical_gap
                near_horizontal = row_ratio >= 0.6 and hgap <= max_horizontal_gap
                
                if not (near_vertical or near_horizontal):
                    continue
                    
                new_tl, new_br = _rect_union(tl_i, br_i, tl_j, br_j)
                new_members = set(members_i)
                new_members.update(members_j)
                blocks[i] = (new_tl, new_br, new_members)
                labels[i] = "table"
                blocks[j] = None
                labels[j] = None
                to_drop.add(j)
                tl_i, br_i, members_i = blocks[i]
                changed = True
                
        if changed:
            compact_blocks = []
            compact_labels = []
            for blk, lbl in zip(blocks, labels):
                if blk is not None and lbl is not None:
                    compact_blocks.append(blk)
                    compact_labels.append(lbl)
            blocks, labels = compact_blocks, compact_labels

    return blocks, labels

# Extractors (simplified)
# Counts the number of non-empty string cells in a given row and column range.
# def _row_string_count(ws: Worksheet, c1: int, c2: int, row: int) -> int:
#     cnt = 0
#     for c in range(c1, c2 + 1):
#         v = ws.cell(row, c).value
#         if isinstance(v, str) and v.strip():
#             cnt += 1
#     return cnt

# Extracts a table block from a worksheet into a structured dictionary.
def extract_table(ws: Worksheet, tl, br) -> Dict[str, Any]:
    r1, c1 = tl; r2, c2 = br
    # Build merged cell map: for each cell in a merged range, assign the value of the top-left cell
    merged_map = {}
    if hasattr(ws, 'merged_cells'):
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            value = ws.cell(min_row, min_col).value
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    merged_map[(rr, cc)] = value
                    
    def _cell_value(row: int, col: int):
        if (row, col) in merged_map:
            return merged_map[(row, col)]
        return ws.cell(row, col).value

    def _row_profile(row: int) -> Tuple[int, int, float, float]:
        values = []
        for col in range(c1, c2 + 1):
            v = _cell_value(row, col)
            if v not in (None, ""):
                values.append(str(v).strip())
        non_empty = len(values)
        if non_empty == 0:
            return 0, 0, 0.0, 0.0
        norm = [v.lower() for v in values]
        unique_count = len(set(norm))
        most_common = max((norm.count(x) for x in set(norm)), default=0)
        unique_ratio = unique_count / non_empty
        most_common_ratio = most_common / non_empty
        return non_empty, unique_count, unique_ratio, most_common_ratio
        
    # --- Generalized: Extract free text above and below the table (all columns) ---
    # (Legacy context_lines logic removed; now handled below)
    # --- Generalized: Extract free text above and below the table (all columns) ---
    # Improved: find the first row with at least 3 non-empty short cells (likely header),
    # and ensure at least 3 following rows have at least 3 non-empty cells (likely data rows)
    min_header_cells = 3
    min_data_rows = 3
    min_data_cells = 3
    header_row = None
    candidate_rows: List[Tuple[int, float]] = []
    
    placeholder_header_patterns = [
        r"coloumn\s*\d+$", # common misspelling that shows up in some sheets
        r"column\s*\d+$",
        r"col\s*\d+$",
        r"unnamed[:\s]*\d+$",
        r"field\s*\d+$",
        r"column_?\d+$",
    ]
    
    def _is_placeholder_header(value: Any) -> bool:
        if not isinstance(value, str):
            return False
        text = value.strip().lower()
        if not text:
            return False
        for pat in placeholder_header_patterns:
            if re.match(pat, text, flags=re.IGNORECASE):
                return True
        return False
        
    def _collect_header_candidates(max_len: Optional[int]) -> List[Tuple[int, float]]:
        candidates: List[Tuple[int, float]] = []
        for rr in range(r1, r2 - min_data_rows + 1):
            values = [_cell_value(rr, c) for c in range(c1, c2 + 1)]
            non_empty = [v for v in values if v not in (None, "")]
            if len(non_empty) < min_header_cells:
                continue
                
            generic_count = sum(1 for v in non_empty if _is_placeholder_header(v))
            generic_ratio = generic_count / len(non_empty) if non_empty else 0.0
            # Skip rows that are mostly placeholder headers like Column1/Column2
            if generic_ratio >= 0.5:
                continue
            if max_len is not None and any(len(str(v)) > max_len for v in non_empty):
                continue
                
            data_like = True
            for offset in range(1, min_data_rows + 1):
                next_values = [_cell_value(rr + offset, c) for c in range(c1, c2 + 1)]
                next_non_empty = [v for v in next_values if v not in (None, "")]
                if len(next_non_empty) < min_data_cells:
                    data_like = False
                    break
            if data_like:
                non_empty_count, _, unique_ratio, most_common_ratio = _row_profile(rr)
                # Penalize rows that contain any placeholder-like headers even if they weren't skipped
                score = (
                    (non_empty_count * 1.0)
                    + (unique_ratio * 10.0)
                    - (most_common_ratio * 5.0)
                    - (generic_ratio * 20.0)
                )
                candidates.append((rr, score))
        return candidates

    # First pass: prefer shorter header labels (legacy behaviour)
    candidate_rows = _collect_header_candidates(max_len=30)
    # Fallback: allow longer header labels (e.g., Zurich sheet has long column names)
    if not candidate_rows:
        candidate_rows = _collect_header_candidates(max_len=120)
        
    if candidate_rows:
        candidate_rows.sort(key=lambda x: (-x[1], x[0]))
        header_row = candidate_rows[0][0]
    else:
        # Last resort: pick the first non-empty row so we don't drop the table entirely
        for rr in range(r1, r2 + 1):
            if any(_cell_value(rr, c) not in (None, "") for c in range(c1, c2 + 1)):
                header_row = rr
                break
                
    # Above-table free text
    context_lines_above = []
    if header_row is not None:
        for row in range(r1, header_row):
            for col in range(c1, c2 + 1):
                val = ws.cell(row, col).value
                if val and str(val).strip():
                    context_lines_above.append(str(val).strip())
                    
    # Below-table free text (up to 3 rows after last data row, but not past r2)
    context_lines_below = []
    last_data_row = None
    if header_row is not None:
        for r in range(header_row + 1, r2 + 1):
            row_values = [ws.cell(r, c).value for c in range(c1, c2 + 1)]
            if any(v not in (None, "") for v in row_values):
                last_data_row = r
        if last_data_row:
            for row in range(last_data_row + 1, min(r2 + 1, last_data_row + 4)):
                for col in range(c1, c2 + 1):
                    val = ws.cell(row, col).value
                    if val and str(val).strip():
                        context_lines_below.append(str(val).strip())
                        
    # Also include first row of block (legacy context)
    context_lines_first_row = []
    for c in range(c1, c2 + 1):
        val = ws.cell(row=r1, column=c).value
        if val and str(val).strip():
            context_lines_first_row.append(str(val).strip())
            
    # Merge all context lines, remove duplicates while preserving order
    context_lines = context_lines_above + context_lines_first_row + context_lines_below
    seen = set()
    context_lines_unique = []
    for line in context_lines:
        if line not in seen:
            context_lines_unique.append(line)
            seen.add(line)
    context = "\n".join(context_lines_unique) if context_lines_unique else None
    
    if header_row is None:
        return {"headers": [], "rows": [], "range": f"{a1(r1, c1)}:{a1(r2, c2)}"}

    # Use only the actual header values, no inference
    headers = []
    for c in range(c1, c2 + 1):
        raw = ws.cell(header_row, c).value
        if (header_row, c) in merged_map:
            raw = merged_map[(header_row, c)]
        text = str(raw).strip() if raw is not None else None
        headers.append({"text": text, "cell": a1(header_row, c)})
        
    # Only keep non-empty header names
    header_indices = [(idx, h) for idx, h in enumerate(headers) if h["text"]]
    header_names = [h["text"] for idx, h in header_indices]
    
    rows = []
    for r in range(header_row + 1, r2 + 1):
        row_obj = []
        for idx, h in header_indices:
            c = c1 + idx
            val = ws.cell(r, c).value
            if (r, c) in merged_map:
                val = merged_map[(r, c)]
            row_obj.append({"header": h["text"], "value": val, "cell": a1(r, c)})
        # Only add row if at least one value is not None or empty
        if any(v["value"] not in (None, "") for v in row_obj):
            rows.append(row_obj)
            
    result = {"type": "table", "range": f"{a1(r1, c1)}:{a1(r2, c2)}", "headers": headers, "rows": rows}
    if context:
        result["context"] = context
    return result

# Extracts key-value pairs from a worksheet block into a structured dictionary.
def extract_kv(ws: Worksheet, tl, br) -> Dict[str, Any]:
    r1, c1 = tl; r2, c2 = br
    items = []
    for r in range(r1, r2 + 1):
        cells = []
        for c in range(c1, c2 + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                cells.append((c, v))
        if not cells:
            continue
        if len(cells) == 1 and isinstance(cells[0][1], str) and ":" in str(cells[0][1]):
            key, val = str(cells[0][1]).split(":", 1)
            items.append({"key": {"value": str(key).strip()}, "value": {"value": str(val).strip()}})
        elif len(cells) >= 2:
            k, v = cells[0], cells[1]
            items.append({"key": {"value": str(k[1]).strip()}, "value": {"value": str(v[1])}})
        else:
            items.append({"key": {"value": ""}, "value": {"value": str(cells[0][1])}})
    return {"type": "key_values", "range": f"{a1(r1, c1)}:{a1(r2, c2)}", "items": items}

# Extracts text lines from a worksheet block into a structured dictionary.
def extract_text(ws: Worksheet, tl, br) -> Dict[str, Any]:
    r1, c1 = tl; r2, c2 = br
    lines = []
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                lines.append({"value": v, "cell": a1(r, c)})
    return {"type": "text_lines", "range": f"{a1(r1, c1)}:{a1(r2, c2)}", "lines": lines}

# --- Sheet profiling for hybrid selection -------------------------
def _sheet_layout_profile(ws: Worksheet) -> Dict[str, Any]:
    """Compute lightweight density and block metrics for a worksheet.
    
    Returns a dict with:
    - non_empty_cells
    - area (bounding box area of non-empty cells)
    - density (non_empty_cells / area)
    - table_blocks
    - table_cells
    - table_cells_ratio (table_cells / non_empty_cells)
    """
    coords = non_empty_coords(ws)
    if not coords:
        return {
            "non_empty_cells": 0,
            "area": 0,
            "density": 0.0,
            "table_blocks": 0,
            "table_cells": 0,
            "table_cells_ratio": 0.0,
        }
        
    raw_blocks = connected_blocks(coords)
    labels = [classify_block(ws, tl, br, members) for (tl, br, members) in raw_blocks]
    merged_blocks, merged_labels = merge_adjacent_blocks(raw_blocks, labels)
    merged_blocks, merged_labels = _merge_tables_with_neighbors(
        merged_blocks, merged_labels, max_vertical_gap=8, max_horizontal_gap=2
    )
    
    # bounding box of non-empty cells
    rs = [r for r, _ in coords]
    cs = [c for _, c in coords]
    area = (max(rs) - min(rs) + 1) * (max(cs) - min(cs) + 1)
    non_empty = len(coords)
    density = (non_empty / area) if area else 0.0
    
    table_cells = 0
    table_blocks = 0
    for blk, lbl in zip(merged_blocks, merged_labels):
        if lbl == "table":
            table_blocks += 1
            table_cells += len(blk[2])
            
    ratio = (table_cells / non_empty) if non_empty else 0.0
    return {
        "non_empty_cells": non_empty,
        "area": area,
        "density": density,
        "table_blocks": table_blocks,
        "table_cells": table_cells,
        "table_cells_ratio": ratio,
    }

# --- Chunking interface --------------------------------------------
# Loads a CSV or Excel file and returns cleaned sheet-level chunks as dictionaries.
def _sheet_level_chunks(file_path: Path) -> List[Dict[str, Any]]:
    file_lower = file_path.name.lower()
    is_csv = file_lower.endswith(".csv")
    is_excel = file_lower.endswith((".xls", ".xlsx", ".xlsm", ".xlsb"))
    if not is_csv and not is_excel:
        raise ValueError(f"Unsupported file type for: {file_path}")
    chunks: List[Dict[str, Any]] = []
    if is_csv:
        df = pd.read_csv(file_path)
        cleaned = clean_df_for_chunk(df)
        chunks.append({"source_type": "csv", "sheet_name": None, "content": cleaned})
        return chunks
    excel_file = pd.ExcelFile(file_path)
    try:
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            cleaned = _clean_df_for_chunk_sheet_level_excel(df)
            chunks.append({"source_type": "excel", "sheet_name": sheet_name, "content": cleaned})
    finally:
        excel_file.close()
    return chunks

def _block_level_chunks(
    file_path: Path,
    max_tokens: int = 2500,
    sheet_filter: Optional[Set[str]] = None,
) -> List[Dict[str, Any]]:
    file_lower = file_path.name.lower()
    is_csv = file_lower.endswith(".csv")
    is_excel = file_lower.endswith((".xls", ".xlsx", ".xlsm", ".xlsb"))
    
    # Fast path for CSV: use pandas and return a single chunk
    if is_csv:
        df = pd.read_csv(file_path)
        cleaned = clean_df_for_chunk(df)
        return [{"source_type": "csv", "sheet_name": None, "content": cleaned}]
        
    if not is_excel:
        raise ValueError(f"Unsupported file type for: {file_path}")
        
    wb = load_workbook(file_path, data_only=True, read_only=False)
    all_chunks: List[Dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            if sheet_filter and ws.title not in sheet_filter:
                continue
            coords = non_empty_coords(ws)
            if not coords:
                continue
            raw_blocks = connected_blocks(coords)
            labels = [classify_block(ws, tl, br, members) for (tl, br, members) in raw_blocks]
            merged_blocks, merged_labels = merge_adjacent_blocks(raw_blocks, labels)
            
            # Tables in these loss runs are sometimes split by subtotal rows or thin text strips.
            # Pull those sparse blocks back into the nearest table when they share columns/rows.
            merged_blocks, merged_labels = _merge_tables_with_neighbors(
                merged_blocks, merged_labels, max_vertical_gap=8, max_horizontal_gap=2
            )
            
            # Collect all block info for context association
            block_info = []
            for (tl, br, members), label in zip(merged_blocks, merged_labels):
                block_info.append({
                    "tl": tl,
                    "br": br,
                    "label": label,
                    "members": members
                })
                
            for idx, block in enumerate(block_info):
                tl, br, label = block["tl"], block["br"], block["label"]
                if label == "table":
                    # Gather context blocks above this table
                    context_blocks = []
                    for prev in block_info:
                        if prev["label"] in ("text_lines", "key_values"):
                            # If the bottom of the context block is above the top of the table
                            if prev["br"][0] < tl[0]:
                                context_blocks.append(prev)
                    # Sort context blocks by their bottom row (top-to-bottom order)
                    context_blocks.sort(key=lambda b: b["br"][0])
                    # Extract text from context blocks
                    context_texts = []
                    for ctx in context_blocks:
                        if ctx["label"] == "text_lines":
                            content = extract_text(ws, ctx["tl"], ctx["br"])
                            lines = [str(line["value"]).strip() for line in content.get("lines", []) if line.get("value")]
                            if lines:
                                context_texts.append("\n".join(lines))
                        elif ctx["label"] == "key_values":
                            content = extract_kv(ws, ctx["tl"], ctx["br"])
                            items = content.get("items", [])
                            kv_lines = []
                            for item in items:
                                k = item.get("key", {}).get("value", "")
                                v = item.get("value", {}).get("value", "")
                                if k:
                                    kv_lines.append(f"{k}: {v}")
                                else:
                                    kv_lines.append(str(v))
                            if kv_lines:
                                context_texts.append("\n".join(kv_lines))
                    context_str = "\n".join(context_texts).strip() if context_texts else None
                    
                    table_content = extract_table(ws, tl, br)
                    merged_context_lines = []
                    for ctx in (context_str, table_content.get("context")):
                        if ctx:
                            for line in str(ctx).splitlines():
                                line = line.strip()
                                if line and line not in merged_context_lines:
                                    merged_context_lines.append(line)
                    merged_context = "\n".join(merged_context_lines) if merged_context_lines else None
                    headers = table_content["headers"]
                    rows = table_content["rows"]
                    header_tokens = len(json_dumps_datetime(headers).split())
                    chunk_rows = []
                    chunk_token_count = header_tokens
                    for row in rows:
                        row_tokens = len(json_dumps_datetime(row).split())
                        if chunk_rows and (chunk_token_count + row_tokens > max_tokens):
                            content_dict = {"type": "table", "headers": headers, "rows": chunk_rows, "range": table_content["range"]}
                            if merged_context:
                                content_dict["context"] = merged_context
                            all_chunks.append({
                                "source_type": "excel",
                                "sheet_name": ws.title,
                                "content": content_dict,
                                "block_type": label,
                                "range": table_content["range"],
                            })
                            chunk_rows = []
                            chunk_token_count = header_tokens
                        chunk_rows.append(row)
                        chunk_token_count += row_tokens
                    if chunk_rows:
                        content_dict = {"type": "table", "headers": headers, "rows": chunk_rows, "range": table_content["range"]}
                        if merged_context:
                            content_dict["context"] = merged_context
                        all_chunks.append({
                            "source_type": "excel",
                            "sheet_name": ws.title,
                            "content": content_dict,
                            "block_type": label,
                            "range": table_content["range"],
                        })
                elif label == "key_values":
                    content = extract_kv(ws, tl, br)
                    all_chunks.append({
                        "source_type": "excel",
                        "sheet_name": ws.title,
                        "content": content,
                        "block_type": label,
                        "range": content.get("range"),
                    })
                else:
                    content = extract_text(ws, tl, br)
                    all_chunks.append({
                        "source_type": "excel",
                        "sheet_name": ws.title,
                        "content": content,
                        "block_type": label,
                        "range": content.get("range"),
                    })
            # --- Floating text extraction: find non-empty cells not in any block ---
            all_block_cells = set()
            for b in merged_blocks:
                all_block_cells.update(b[2])
            floating_cells = coords - all_block_cells
            for (r, c) in sorted(floating_cells):
                val = ws.cell(r, c).value
                if val and str(val).strip():
                    all_chunks.append({
                        "source_type": "excel",
                        "sheet_name": ws.title,
                        "content": {
                            "type": "floating_text",
                            "range": f"{a1(r, c)}:{a1(r, c)}",
                            "lines": [str(val).strip()]
                        },
                        "block_type": "floating_text",
                        "range": f"{a1(r, c)}:{a1(r, c)}",
                    })
    finally:
        wb.close()
    return all_chunks

# Estimates the number of tokens in a JSON-serializable object for budgeting chunk size.
def _estimate_tokens(obj: Any) -> int:
    """
    Estimates the number of tokens in a JSON-serializable object using tiktoken.
    Falls back to word-based estimation if tiktoken is unavailable or fails.
    """
    try:
        encoding = tiktoken.get_encoding("o200k_base")
        text = json_dumps_datetime(obj)
        tokens = encoding.encode(text)
        return len(tokens)
    except Exception:
        import re as _re
        text = json_dumps_datetime(obj)
        words = _re.findall(r"\b\w+\b", text)
        short = sum(1 for w in words if len(w) <= 4)
        long = sum(1 for w in words if len(w) > 4)
        return short + int(long * (100 / 75))

# Splits a dictionary of rows into multiple chunks based on a token budget.
# Each chunk contains as many rows as possible without exceeding the max token limit.
def _split_rows_by_token_budget(
    rows_dict: Dict[str, Dict[str, Any]],
    source_type: str,
    sheet_name: Optional[str],
    max_tokens: int,
) -> List[Dict[str, Any]]:
    buckets: List[Dict[str, Any]] = []
    current: Dict[str, Dict[str, Any]] = {}
    
    def _row_sort_key(key: str) -> Tuple[int, str]:
        match = re.search(r"\b(\d+)\b", key)
        return (int(match.group(1)) if match else 10**9, key)
        
    for row_key in sorted(rows_dict.keys(), key=_row_sort_key):
        candidate = dict(current)
        candidate[row_key] = rows_dict[row_key]
        if current and _estimate_tokens(candidate) > max_tokens:
            buckets.append({"source_type": source_type, "sheet_name": sheet_name, "content": current})
            current = {row_key: rows_dict[row_key]}
        else:
            current = candidate
    if current:
        buckets.append({"source_type": source_type, "sheet_name": sheet_name, "content": current})
    return buckets

# --- LLM JSON formatting utility ----------------------------------
def format_chunk_for_llm_json(chunk: dict) -> str:
    """
    Formats a chunk (as produced by get_excel_csv_chunks_advanced) as a JSON string for LLM input.
    Includes context, headers, and rows in a compact, LLM-friendly structure.
    """
    content = chunk.get("content", {})
    # Minimal JSON for table chunks: {context, headers, rows}
    if content.get("type") == "table":
        out = {}
        # Prepend context as plain text if present
        if "context" in content and content["context"]:
            out["context"] = _sanitize_llm_text(content["context"])
        out["headers"] = [_sanitize_llm_text(h["text"]) for h in content.get("headers", [])]
        out["rows"] = [
            [_sanitize_llm_text(cell.get("value", None)) for cell in row]
            for row in content.get("rows", [])
        ]
        return json_dumps_datetime(out, separators=(",", ":"))
    # For other types, fallback to previous minimal output
    elif content.get("type") == "key_values":
        return json_dumps_datetime({
            "key_values": [
                {
                    "key": _sanitize_llm_text(item["key"]["value"]),
                    "value": _sanitize_llm_text(item["value"]["value"]),
                }
                for item in content.get("items", [])
            ]
        }, separators=(",", ":"))
    elif content.get("type") == "text_lines":
        return json_dumps_datetime({
            "lines": [_sanitize_llm_text(line["value"]) for line in content.get("lines", [])]
        }, separators=(",", ":"))
    elif content.get("type") == "floating_text":
        return json_dumps_datetime({
            "floating_text": [_sanitize_llm_text(v) for v in content.get("lines", [])]
        }, separators=(",", ":"))
    else:
        if isinstance(content, dict):
            sanitized = {}
            for k, v in content.items():
                if isinstance(v, str):
                    sanitized[k] = _sanitize_llm_text(v)
                elif isinstance(v, list):
                    sanitized[k] = [_sanitize_llm_text(i) for i in v]
                else:
                    sanitized[k] = v
            return json_dumps_datetime(sanitized, separators=(",", ":"))
        return json_dumps_datetime(content, separators=(",", ":"))

# Main interface: loads and chunks an Excel/CSV file, optionally using block extraction and token budgeting.
# Returns a list of numbered chunks.
def get_excel_csv_chunks_advanced(
    file_path: str,
    hybrid_approach: bool = True,
    max_tokens: int = 2500,
) -> List[Dict[str, Any]]:
    """
    Hybrid sheet/block chunker.
    
    - If `hybrid_approach` is False, return the original sheet-level cleaned chunks
      (no token budgeting or block detection).
    - If `hybrid_approach` is True, first build sheet-level chunks. For each sheet,
      if its estimated token count exceeds `max_tokens`, fall back to block-level
      chunking for that sheet only; otherwise keep the simple sheet chunk.
    """
    path = Path(file_path)
    print(f"[CHUNKING] {path.name} -> hybrid={hybrid_approach} | max_tokens={max_tokens}")
    prepared_path, cleanup_paths = _ensure_chunkable_copy(path)
    try:
        file_lower = prepared_path.name.lower()
        is_csv = file_lower.endswith(".csv")
        is_excel = file_lower.endswith((".xls", ".xlsx", ".xlsm", ".xlsb"))
        if not is_csv and not is_excel:
            raise ValueError(f"Unsupported file type for: {file_path}")
            
        if not hybrid_approach:
            chunks = _sheet_level_chunks(prepared_path)
            for ch in chunks:
                src = ch.get("source_type")
                sheet = ch.get("sheet_name") or "<csv>"
                estimated = _estimate_tokens(ch.get("content", {}))
                print(
                    f"[CHUNKING] sheet '{sheet}': method=sheet-level (hybrid_approach=False; estimated={estimated}; max_tokens={max_tokens})"
                )
        else:
            sheet_chunks = _sheet_level_chunks(prepared_path)
            
            if is_csv:
                # CSV behaves as a single sheet; if it exceeds max_tokens, split by row budget (no block-level).
                if not sheet_chunks:
                    chunks = []
                else:
                    only_chunk = sheet_chunks[0]
                    content = only_chunk.get("content", {})
                    estimated = _estimate_tokens(content)
                    if estimated > max_tokens and isinstance(content, dict):
                        chunks = _split_rows_by_token_budget(
                            content,
                            source_type="csv",
                            sheet_name=None,
                            max_tokens=max_tokens,
                        )
                        print(
                            f"[CHUNKING] sheet '<csv>': method=sheet-level-split (estimated={estimated}; max_tokens={max_tokens})"
                        )
                    else:
                        chunks = [only_chunk]
                        print(
                            f"[CHUNKING] sheet '<csv>': method=sheet-level (estimated={estimated}; max_tokens={max_tokens})"
                        )
            else:
                # Decide per-sheet whether block-level chunking is needed
                block_sheet_names: Set[str] = set()
                est_by_sheet: Dict[str, int] = {}
                decision_by_sheet: Dict[str, str] = {}
                for ch in sheet_chunks:
                    content = ch.get("content", {})
                    estimated = _estimate_tokens(content)
                    sheet_name = ch.get("sheet_name")
                    if sheet_name is None:
                        continue
                    est_by_sheet[sheet_name] = estimated
                    if estimated > max_tokens:
                        block_sheet_names.add(sheet_name)
                        decision_by_sheet[sheet_name] = "block-level"
                    else:
                        decision_by_sheet[sheet_name] = "sheet-level"
                        
                block_chunks: List[Dict[str, Any]] = []
                if block_sheet_names:
                    block_chunks = _block_level_chunks(
                        prepared_path,
                        max_tokens=max_tokens,
                        sheet_filter=block_sheet_names,
                    )
                    
                block_by_sheet: Dict[str, List[Dict[str, Any]]] = {}
                for ch in block_chunks:
                    block_by_sheet.setdefault(ch.get("sheet_name"), []).append(ch)
                    
                chunks = []
                for ch in sheet_chunks:
                    sheet_name = ch.get("sheet_name")
                    if sheet_name in block_sheet_names:
                        replacement = block_by_sheet.get(sheet_name)
                        if replacement:
                            chunks.extend(replacement)
                            print(
                                f"[CHUNKING] sheet '{sheet_name}': method=block-level (estimated={est_by_sheet.get(sheet_name, 'n/a')}; max_tokens={max_tokens})"
                            )
                        else:
                            # Fallback: keep sheet-level chunk if block extraction produced nothing
                            chunks.append(ch)
                            print(
                                f"[CHUNKING] sheet '{sheet_name}': method=sheet-level-fallback (estimated={est_by_sheet.get(sheet_name, 'n/a')}; max_tokens={max_tokens})"
                            )
                    else:
                        chunks.append(ch)
                        print(
                            f"[CHUNKING] sheet '{sheet_name}': method=sheet-level (estimated={est_by_sheet.get(sheet_name, 'n/a')}; max_tokens={max_tokens})"
                        )
    finally:
        for temp_path in cleanup_paths:
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception:
                # Best-effort cleanup; ignore failures to avoid masking upstream errors.
                pass
                
    # number them
    result: List[Dict[str, Any]] = []
    csv_chunk_count = sum(1 for ch in chunks if ch.get("source_type") == "csv")
    for idx, chunk in enumerate(chunks, start=1):
        c = dict(chunk)
        c["chunk_number"] = idx
        sheet = c.get("sheet_name") or "<csv>"
        estimated = _estimate_tokens(c.get("content", {}))
        if c.get("source_type") == "csv":
            method = "sheet-level-split" if csv_chunk_count > 1 else "sheet-level"
        elif c.get("block_type"):
            method = "block-level"
        else:
            method = "sheet-level"
        print(
            f"[CHUNKING] chunk {idx} (sheet '{sheet}'): method={method} (estimated={estimated}; max_tokens={max_tokens})"
        )
        result.append(c)
    return result

# --- Save JSON to company folder ---
def save_chunks_to_company_folder(chunks, company_name, base_output_dir, filename="extracted_chunks.json"):
    """
    Saves the given chunks (list of dicts) as a JSON file in the specified company's folder under base_output_dir.
    Creates the directory if it does not exist.
    
    Args:
        chunks: List of chunk dicts (output of get_excel_csv_chunks_advanced)
        company_name: Name of the company (str)
        base_output_dir: Base directory where company folders are located (str or Path)
        filename: Name of the JSON file to write (default: 'extracted_chunks.json')
        
    Returns:
        The path to the saved JSON file (as a Path object)
    """
    # from pathlib import Path
    # import json
    company_dir = Path(base_output_dir) / company_name
    _ensure_dir(company_dir)
    
    output_path = company_dir / filename
    # Handle long output paths on Windows (e.g., long company/file names)
    target_path = output_path
    if os.name == "nt":
        resolved = output_path.resolve()
        if len(str(resolved)) >= 240:
            target_path = _as_extended_path(resolved)
    with open(target_path, "w", encoding="utf-8") as f:
        json.dump(chunks, f, indent=2, ensure_ascii=False, default=str)
    return output_path

# Convenience CLI for quick runs
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Hybrid Excel/CSV chunking")
    parser.add_argument("file", help="Path to Excel or CSV file")
    parser.add_argument("--no-hybrid", action="store_true", help="Disable hybrid logic (always sheet-level chunks)")
    parser.add_argument("--json-out", help="Optional path to write JSON output")
    parser.add_argument("--llm-json", action="store_true", help="Print each chunk as LLM-ready JSON string (for inspection or LLM input)")
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=2500,
        help="Token budget per sheet (and per block-chunk when needed)",
    )
    
    args = parser.parse_args()
    
    chunks = get_excel_csv_chunks_advanced(
        args.file,
        hybrid_approach=not args.no_hybrid,
        max_tokens=args.max_tokens,
    )
    
    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            f.write(json_dumps_datetime(chunks, indent=2))
    elif args.llm_json:
        for chunk in chunks:
            print(format_chunk_for_llm_json(chunk))
    else:
        print(json_dumps_datetime(chunks, indent=2))
